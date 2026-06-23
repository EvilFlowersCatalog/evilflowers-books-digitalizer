"""Build a metadata catalog *draft* for librarians to fill in.

We can derive a few fields from the directory name (the exact join key, the
faculty, an ISBN for the ~80% of dirs that are ISBNs, and a guessed title for
the title-slug dirs). The rest is left blank for a librarian to complete from
the physical book / OPAC. The output is a styled ``.xlsx`` with a frozen,
colour-coded header (grey = pre-filled, do not edit; yellow = please fill) and
an instructions sheet in Slovak.

The header names here are the single source of truth: ``configs/pipeline.toml``
``[metadata.columns]`` maps the model fields onto these same headers (a test in
``tests/test_draft.py`` keeps the two in sync).
"""

from __future__ import annotations

import re
from dataclasses import dataclass

from evilflowers_books_digitalizer.metadata.catalog import deslug, extract_dir_id

# Slovak column headers, in display order. The first four are pre-filled by us.
HEADERS: dict[str, str] = {
    "directory_id": "directory_id",  # exact dir name — JOIN KEY, do not edit
    "faculty": "fakulta",
    "n_pages": "počet_strán",
    "isbn": "ISBN",
    "title": "názov",
    "subtitle": "podnázov",
    "authors": "autori",
    "year": "rok_vydania",
    "publisher": "vydavateľ",
    "place": "miesto_vydania",
    "edition": "vydanie",
    "language": "jazyk",
    "catalog": "cieľový_katalog",  # optional: override the import target catalog
    "notes": "poznámka",
    "source_meta": "zdroj_metadat",  # provenance of any auto-filled fields
}

#: Columns we pre-fill (greyed, informative); the rest are for the librarian.
PREFILLED = ("directory_id", "faculty", "n_pages", "isbn", "title", "source_meta")

#: Fields an ISBN lookup may pre-fill (when the dir name is a valid ISBN).
ENRICHED_FIELDS = (
    "title", "subtitle", "authors", "year", "publisher", "place", "edition", "language",
)

#: Per-column guidance shown on the instructions sheet.
INSTRUCTIONS: dict[str, str] = {
    "directory_id": "Identifikátor adresára na úložisku. NEUPRAVOVAŤ — slúži na spárovanie.",
    "fakulta": "Fakulta (predvyplnené).",
    "počet_strán": "Počet naskenovaných snímok (predvyplnené, informatívne).",
    "ISBN": "ISBN, ak je známe. Predvyplnené tam, kde sa dalo odvodiť z názvu adresára — prosím overte.",
    "názov": "Názov publikácie. Pri ISBN adresároch je prázdny — doplňte. Pri slovných je to odhad — opravte.",
    "podnázov": "Podnázov / časť názvu (nepovinné).",
    "autori": "Autori oddelení bodkočiarkou, napr. „Novák Ján; Malá Eva“.",
    "rok_vydania": "Rok vydania (štvorciferný).",
    "vydavateľ": "Vydavateľ.",
    "miesto_vydania": "Miesto vydania (nepovinné).",
    "vydanie": "Poradie vydania, napr. „2. preprac. vyd.“ (nepovinné).",
    "jazyk": "Jazyk: sk / en / cs / de / ru (nepovinné — inak sa zistí automaticky).",
    "cieľový_katalog": "Cieľový katalóg EvilFlowers (nepovinné — prázdne = predvolený katalóg z konfigurácie).",
    "poznámka": "Ľubovoľná poznámka (nepovinné).",
    "zdroj_metadat": "Zdroj predvyplnených údajov z ISBN (napr. Open Library) — prosím overte. Prázdne = vypĺňa knihovník.",
}

_ISBN13 = re.compile(r"^\d{13}$")
_ISBN10 = re.compile(r"^\d{9}[\dXx]$")
_ISBN10_X = re.compile(r"^\d{9}_[Xx]$")


@dataclass
class DraftBook:
    """One book to seed a draft row."""

    source: str
    book_id: str  # exact directory name
    faculty: str
    n_pages: int | None = None


def derive_isbn(book_id: str) -> str:
    """Return an ISBN derived from the directory id, or '' if it isn't one."""
    token = extract_dir_id(book_id)
    if _ISBN13.match(token) or _ISBN10.match(token):
        return token.upper()
    if _ISBN10_X.match(token):
        return token.replace("_", "").upper()  # 807095020_X -> 807095020X
    return ""


def guess_title(book_id: str) -> str:
    """De-slugged title for slug dirs; '' for ISBN/numeric dirs (let librarian fill)."""
    token = extract_dir_id(book_id)
    if re.fullmatch(r"[0-9_Xx]+", token):
        return ""  # an ISBN/OPAC number is not a title
    return deslug(book_id)


def build_draft_rows(books: list[DraftBook], enricher: object | None = None) -> list[dict[str, object]]:
    """Seed one ordered dict per book, keyed by the Slovak headers.

    ``enricher`` (anything with ``.lookup(isbn) -> dict``, e.g.
    :class:`~.isbn_lookup.IsbnEnricher`) optionally pre-fills bibliographic
    fields for books whose directory name is a valid ISBN. Lookups are
    best-effort: a miss just leaves the cells blank for the librarian.
    """
    rows: list[dict[str, object]] = []
    for b in books:
        isbn = derive_isbn(b.book_id)
        seed: dict[str, object] = {
            "directory_id": b.book_id,
            "faculty": b.faculty,
            "n_pages": b.n_pages,
            "isbn": isbn,
            "title": guess_title(b.book_id),
        }
        if enricher is not None and isbn:
            found = enricher.lookup(isbn)  # type: ignore[attr-defined]
            if found:
                _apply_enrichment(seed, found)
        rows.append({HEADERS[k]: seed.get(k, "") for k in HEADERS})
    return rows


def _apply_enrichment(seed: dict[str, object], found: dict) -> None:
    """Merge ISBN-lookup fields into a row seed (real title wins over the guess)."""
    if found.get("title"):
        seed["title"] = found["title"]
    if found.get("subtitle"):
        seed["subtitle"] = found["subtitle"]
    if found.get("authors"):
        seed["authors"] = "; ".join(found["authors"])
    if found.get("year"):
        seed["year"] = found["year"]
    if found.get("publisher"):
        seed["publisher"] = found["publisher"]
    if found.get("place"):
        seed["place"] = found["place"]
    if found.get("edition"):
        seed["edition"] = found["edition"]
    if found.get("language"):
        seed["language"] = found["language"]
    seed["source_meta"] = found.get("_source", "")


def write_draft_xlsx(books: list[DraftBook], path, enricher: object | None = None) -> "object":
    """Write a styled, librarian-ready draft workbook; returns the Path.

    Pass an ``enricher`` to auto-fill metadata for ISBN dirs (see
    :func:`build_draft_rows`).
    """
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = build_draft_rows(books, enricher=enricher)
    headers = list(HEADERS.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "katalog"

    grey = PatternFill("solid", fgColor="D9D9D9")  # pre-filled by us
    yellow = PatternFill("solid", fgColor="FFF2CC")  # empty — please fill
    green = PatternFill("solid", fgColor="E2EFDA")  # auto-filled from ISBN — verify
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF")
    prefilled_headers = {HEADERS[k] for k in PREFILLED}
    enriched_headers = {HEADERS[k] for k in ENRICHED_FIELDS}

    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 34 if header in ("názov", "autori") else 16

    for r in rows:
        ws.append([r[h] for h in headers])

    # tint the body: grey = pre-filled column; for librarian columns, green when
    # auto-filled from an ISBN lookup (verify), yellow when empty (fill).
    for col_idx, header in enumerate(headers, start=1):
        for row_idx, r in enumerate(rows, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if header in prefilled_headers:
                cell.fill = grey
            elif header in enriched_headers and r.get(header):
                cell.fill = green
            else:
                cell.fill = yellow

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # instructions sheet
    info = wb.create_sheet("pokyny")
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 90
    title_cell = info.cell(row=1, column=1, value="Pokyny pre vyplnenie katalógu")
    title_cell.font = Font(bold=True, size=14)
    info.cell(
        row=2,
        column=1,
        value=(
            "Sivé bunky sú predvyplnené (neupravujte directory_id). "
            "Zelené bunky sú automaticky doplnené z ISBN — prosím overte. "
            "Žlté bunky prosím doplňte."
        ),
    )
    info.cell(row=4, column=1, value="stĺpec").font = Font(bold=True)
    info.cell(row=4, column=2, value="význam").font = Font(bold=True)
    for i, (header, text) in enumerate(INSTRUCTIONS.items(), start=5):
        info.cell(row=i, column=1, value=header).font = Font(bold=True)
        info.cell(row=i, column=2, value=text).alignment = Alignment(wrap_text=True)

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
