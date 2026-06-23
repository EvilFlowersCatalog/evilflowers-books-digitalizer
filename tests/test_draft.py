"""Draft-builder helpers + config/headers sync."""

from __future__ import annotations

import tomllib
from pathlib import Path

from openpyxl import load_workbook

from evilflowers_books_digitalizer.metadata import MetadataCatalog
from evilflowers_books_digitalizer.metadata.draft import (
    HEADERS,
    DraftBook,
    build_draft_rows,
    derive_isbn,
    guess_title,
    write_draft_xlsx,
)

CONFIG = Path(__file__).resolve().parent.parent / "configs" / "pipeline.toml"


def test_derive_isbn_handles_all_dir_formats():
    assert derive_isbn("CVI_OPACID_FA_9788022750462") == "9788022750462"  # isbn13
    assert derive_isbn("CVI_OPACID_FA_0415128269") == "0415128269"  # isbn10
    assert derive_isbn("CVI_OPACID_FA_807095020_X") == "807095020X"  # isbn10 + X
    assert derive_isbn("CVI_OPACID_FA_Betonove_konstrukcie") == ""  # slug -> no isbn


def test_guess_title_only_for_slugs():
    assert guess_title("CVI_OPACID_FA_Betonove_konstrukcie") == "Betonove Konstrukcie"
    assert guess_title("CVI_OPACID_FA_9788022750462") == ""  # number is not a title


def test_build_rows_seed_prefilled_columns():
    rows = build_draft_rows(
        [DraftBook("fad", "CVI_OPACID_FA_0415128269", "FAD", n_pages=120)]
    )
    row = rows[0]
    assert row["directory_id"] == "CVI_OPACID_FA_0415128269"
    assert row["fakulta"] == "FAD"
    assert row["počet_strán"] == 120
    assert row["ISBN"] == "0415128269"


def test_config_columns_match_draft_headers():
    cfg = tomllib.load(CONFIG.open("rb"))["metadata"]
    for header in cfg["columns"].values():
        assert header in HEADERS.values(), f"config column {header!r} not in draft HEADERS"
    assert cfg["sheet"] == "katalog"


def test_draft_roundtrips_through_catalog(tmp_path):
    books = [
        DraftBook("fad", "CVI_OPACID_FA_0415128269", "FAD", 100),
        DraftBook("svf", "CVI_OPACID_SVF_Betonove_konstrukcie", "SVF", 50),
    ]
    out = write_draft_xlsx(books, tmp_path / "catalog.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    assert {"katalog", "pokyny"} <= set(wb.sheetnames)

    cfg = tomllib.load(CONFIG.open("rb"))["metadata"]
    cat = MetadataCatalog.from_excel(
        out, sheet=cfg["sheet"], columns=cfg["columns"], key_field=cfg["key_field"]
    )
    report = cat.match_report([b.book_id for b in books])
    assert report["missed"] == 0  # exact directory join matches both
